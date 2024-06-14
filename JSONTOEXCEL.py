import json
from openpyxl import Workbook

# Load the JSON data from a file
with open('progress.json', 'r') as file:
    data = json.load(file)

# Create a new workbook
wb = Workbook()

# Function to create a sheet for each high-level dictionary
def create_sheet(wb, sheet_name, data_dict):
    sheet = wb.create_sheet(title=sheet_name)

    # Get the keys from the first dictionary to create headers
    headers = set()
    for key, entries in data_dict.items():
        if isinstance(entries, list):
            for entry in entries:
                headers.update(entry.keys())
        elif isinstance(entries, dict):
            headers.update(entries.keys())

    headers = list(headers)
    headers.insert(0, sheet_name[:-1] + ' Key')  # Add the key header

    # Write headers to the sheet
    for col, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col, value=header)

    # Write data to the sheet
    row = 2
    for key, entries in data_dict.items():
        if isinstance(entries, list):
            for entry in entries:
                sheet.cell(row=row, column=1, value=key)  # Write the key
                for col, header in enumerate(headers[1:], 2):
                    sheet.cell(row=row, column=col, value=entry.get(header))
                row += 1
        elif isinstance(entries, dict):
            sheet.cell(row=row, column=1, value=key)  # Write the key
            for col, header in enumerate(headers[1:], 2):
                sheet.cell(row=row, column=col, value=entries.get(header))
            row += 1

# Create a sheet for each high-level dictionary
for key in data:
    create_sheet(wb, key, data[key])

# Remove the default sheet created by openpyxl
default_sheet = wb['Sheet']
wb.remove(default_sheet)

# Save the workbook
wb.save('output.xlsx')
