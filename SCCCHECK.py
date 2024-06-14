import openpyxl
import re
import argparse
import os
from datetime import datetime

def read_excel(file_path): # Load an Excel workbook from the specified file path
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        return workbook
    except PermissionError:
        print(f"Permission denied: Unable to access '{file_path}'.") # handles files that are open somewhere else
        return None

def find_value_with_regex(sheet, pattern, max_diff=5, max_rows=150, max_cols=50): # Find the first cell that matches a regex pattern in the specified Excel sheet
    if isinstance(pattern, str):
        compiled_pattern = re.compile(pattern, re.IGNORECASE)
    else:
        compiled_pattern = pattern  

    for row in sheet.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True):
        for cell in row:
            if cell and regex_match_with_diff(compiled_pattern, str(cell), max_diff):
                return str(cell)
    return False

def regex_match_with_diff(pattern, text, max_diff): # Match text to a regex pattern, allowing for flexible slicing within max_diff
    for i in range(len(text) - max_diff + 1):
        for j in range(i + 1, len(text) + 1):
            if pattern.fullmatch(text[i:j]):
                return True
    return False

def find_most_recent_date(sheet, max_rows=150): # Find most recent review date 
    latest_date = None
    for row in sheet.iter_rows(min_row=1, max_row=max_rows, values_only=True):
        for cell in row:
            if isinstance(cell, datetime): # grabs all that are dates and compares for most recent
                if not latest_date or cell > latest_date:
                    latest_date = cell
    return latest_date

def check_column_presence(workbook, column_name, max_cols=50): # Check if a column with certain name exists on all sheets except first
    for sheet_name in workbook.sheetnames[1:]:
        sheet = workbook[sheet_name]
        for column in sheet.iter_cols(min_row=1, max_row=1, max_col=max_cols, values_only=True):
            if column[0] and column_name.lower() in column[0].lower():
                return True
    return False

def check_reviewed_within_days(last_review_date, days=180): # Check if the last review date is within the specified number of days from today
    if last_review_date:
        return (datetime.now() - last_review_date).days <= days
    return False

def process_scc_file(file_path): # Run checks on SCC file
    print(f'Performing SCC checks on {file_path}')
    workbook = read_excel(file_path)

    if workbook is None:
        print(f"Skipping file due to error: {file_path}") # error handling
        return {}

    first_sheet = workbook[workbook.sheetnames[0]]
    scm_pattern = re.compile(r'SCM\d+', re.IGNORECASE)
    guidance_pattern = re.compile(r'SCC Guidance Source', re.IGNORECASE)
    policy_procedure_pattern = re.compile(r'SCC Policy and Procedures Source', re.IGNORECASE)
    system_scope_pattern = re.compile(r'SCC System Scope', re.IGNORECASE)

    # Extract SCC name from the file path and remove extension
    scc_name = os.path.splitext(os.path.basename(file_path))[0]
    
    # Extract version number from the SCC name
    version_match = re.search(r'_(\d{2})$', scc_name)
    version = version_match.group(1) if version_match else None
    
    # Remove the trailing "_**" from the SCC name
    scc_name = re.sub(r'_\d{2}$', '', scc_name)

    # load information into dictionary
    scc_info = {
        'SCC': scc_name,
        'Version': version,
        'SCM Name': find_value_with_regex(first_sheet, scm_pattern, 0),
        'Last Review Date': find_most_recent_date(first_sheet),
        'SCC Guidance source presence': bool(find_value_with_regex(first_sheet, guidance_pattern, 5)),
        'SCC Policy and Procedure presence': bool(find_value_with_regex(first_sheet, policy_procedure_pattern, 5)),
        'Exception column presence': check_column_presence(workbook, 'exception'),
        'Deviation column presence': check_column_presence(workbook, 'deviation'),
        'TLA column presence': check_column_presence(workbook, 'TLA'),
        'Compliance method column presence': check_column_presence(workbook, 'method'),
        'WPS config sup doc presence': check_column_presence(workbook, 'documentation'),
        'Reviewed within 180 days': check_reviewed_within_days(find_most_recent_date(first_sheet)),
        'SCC System Scope Presence': bool(find_value_with_regex(first_sheet, system_scope_pattern, 3)),
        'Directory built': False
    }

    return scc_info

def main(): 
    parser = argparse.ArgumentParser(description='Analyze SCC files.')
    parser.add_argument('file_path', type=str, help='Path to the Excel file')
    args = parser.parse_args()

    if not os.path.isfile(args.file_path):
        print(f"File not found: {args.file_path}")
        return

    scc_info = process_scc_file(args.file_path)
    for key, value in scc_info.items():
        print(f"{key}: {value}")

if __name__ == "__main__":
    main()
