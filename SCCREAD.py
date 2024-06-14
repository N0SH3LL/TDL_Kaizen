import openpyxl
import re
import argparse
import os

def read_excel(file_path): #error handling for file opening
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        return workbook
    except PermissionError:
        print(f"Permission denied: Unable to access '{file_path}'. The file may be open in another program.")
    except ValueError as e:
        print(f"ValueError: Unable to read workbook: {file_path}. Error: {e}")
    except Exception as e:
        print(f"Unexpected error while reading '{file_path}': {e}")

    return None

def find_unique_values(sheet, column_index): #read a column and separate unique values
    unique_values = set()
    for row in sheet.iter_rows(min_row=2, max_row=1000, min_col=column_index, max_col=column_index, values_only=True):
        cell_value = row[0]
        if cell_value:
            unique_values.add(str(cell_value))
    return unique_values

def split_documentation_text(text): #split documents within cells separated by 4 or more spaces
    return re.split(r'\s{2,}|\r?\n', text)

def process_excel_file(file_path): # just helps process_workbook, could be refined
    print(f'Processing {file_path}')
    workbook = read_excel(file_path)

    if workbook is None:
        print(f"Skipping file due to error: {file_path}")
        return {}, {}, {}, {}  # Return empty dictionaries if the workbook couldn't be opened

    bper_dict, doc_dict, attestation_dict, method_dict = process_workbook(workbook, file_path)
    
    return bper_dict, doc_dict, attestation_dict, method_dict

def process_workbook(workbook, file_path): # grabs all the information from the SCC excel file and puts it into dictionaries
    bper_dict = {}
    doc_dict = {}
    attestation_dict = {}
    method_dict = {}

    attestation_pattern = re.compile(r'(?<!\w)\d{6}(?!\w)') # matching for attestations

    # Extract SCC name from the file path and remove extension and trailing "_**"
    scc_name = os.path.splitext(os.path.basename(file_path))[0]
    scc_name = re.sub(r'_\d{2}$', '', scc_name).strip()

    for sheet_name in workbook.sheetnames[1:]: # check the headers on every tab except the first
        sheet = workbook[sheet_name]
        col_index = 1  # Start from the first column
        for column in sheet.iter_cols(min_row=1, max_row=1, max_col=50, values_only=True): # look through every column
            header = column[0]
            if not header:
                col_index += 1
                continue
            header = str(header).lower() # grab the header

            if 'exception' in header or 'deviation' in header:  # check header for this text and grab all BPER names
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=col_index, max_col=col_index, values_only=True):
                    cell_value = row[0]
                    if cell_value:
                        cell_value_str = str(cell_value)
                        bper_pattern = re.compile(r'BPER\d{7}')  # Regular expression pattern for BPER000****
                        bper_values = bper_pattern.findall(cell_value_str)  # Find all BPER values in the cell
                        
                        for bper_value in bper_values:
                            bper_value = bper_value.strip()
                            if bper_value:
                                bper_dict[bper_value] = {  # BPER dict initialization
                                    'SCC': scc_name,
                                    'BPER name': bper_value,
                                    'Approval Status': '',
                                    'Valid to': '',
                                    'Gathered': False,
                                    'TLA': ''
                                }

            elif 'documentation' in header: # check header for this text and grab all document names
                for cell in sheet.iter_rows(min_row=2, max_row=1000, min_col=col_index, max_col=col_index, values_only=True):
                    if cell[0] is not None:
                        cell_value_str = str(cell[0])
                        doc_names = re.split(r'\s{2,}|\r?\n', cell_value_str) # create a list with all the document names
                        for doc_name in doc_names:
                            doc_name_for_comparison = doc_name.replace('\n', '').strip().upper()  # Normalize case for comparison

                            if doc_name_for_comparison in ['NA', 'N/A', 'NO', 'NONE']: # handles placeholders
                                continue

                            attestation_match = attestation_pattern.search(doc_name) # attestation logic placed here because they end up in this column
                            if attestation_match:
                                attestation_num = attestation_match.group() # adds all attestations that match the pattern
                                if attestation_num not in attestation_dict:
                                    attestation_dict[attestation_num] = { # attestation dict initialization
                                        'SCC': scc_name,
                                        'Attestation num': attestation_num,
                                        'Gathered': False,
                                        'Approval Status': '',
                                        'Valid to': ''
                                    }
                            else:
                                doc_name_final = re.sub(r'\b\d{6}\b', '', doc_name).strip()
                                if doc_name_final and doc_name_final not in doc_dict:
                                    doc_dict[doc_name_final] = { # supporting document ditc initialization
                                        'SCC': scc_name,
                                        'Doc name': doc_name_final,
                                        'Version':'',
                                        'Last update': '',
                                        'Gathered': False                                        
                                    }

            elif 'method' in header: # finds the method column 
                for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_row=1000, min_col=col_index, max_col=col_index, values_only=True), start=2):
                    if row[0] is not None: # grabs them all unless its blank
                        method = str(row[0])
                        stig_id_cell = sheet.cell(row=row_index, column=1)
                        stig_id = str(stig_id_cell.value)
                        method_dict[stig_id] = { # check method dict initialization
                            'SCC': scc_name,
                            'STIG ID': stig_id,
                            'Evidence Method': method,
                            'compliant': '',
                            'Gathered': False
                        }
            col_index += 1

    return bper_dict, doc_dict, attestation_dict, method_dict

def main(): 
    parser = argparse.ArgumentParser(description='Grabs BPERs and docs from an SCC.')
    parser.add_argument('file_path', type=str, help='Path to the Excel file')
    args = parser.parse_args()

    if not os.path.isfile(args.file_path):
        print(f"File not found: {args.file_path}")
        return

    bper_dict, doc_dict, attestation_dict, method_dict = process_excel_file(args.file_path)

    # Printing BPER Names
    print("BPER Names:")
    for bper in bper_dict:
        print(bper)
    
    # Printing Document Names
    print("\nDocument Names:")
    for doc in doc_dict:
        print(doc)
    
    # Printing Attestation Nums
    print("\nAttestation Numbers:")
    for attestation in attestation_dict:
        print(attestation)
    
    # Printing STIG ID and compliance methods
    print("\nCompliance Methods:")
    for stig_id, method_info in method_dict.items():
        print(f"STIG ID: {method_info['STIG ID']}, Method: {method_info['Method']}")

if __name__ == "__main__":
    main()
